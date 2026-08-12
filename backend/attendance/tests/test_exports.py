import csv
import io

from django.urls import reverse
from django.contrib.auth import get_user_model
from rest_framework.test import APITestCase
from rest_framework import status
from rest_framework.authtoken.models import Token
from accounts.models import StudentProfile
from attendance.models import AttendanceSession, Attendance

User = get_user_model()


class ExportTest(APITestCase):
    def setUp(self):
        self.teacher = User.objects.create_user(username="prof", password="pw12345678", role=User.ROLE_TEACHER)
        self.teacher_token = Token.objects.create(user=self.teacher)
        self.student_token = Token.objects.create(
            user=User.objects.create_user(username="stud0", password="pw12345678", role=User.ROLE_STUDENT)
        )
        self.s1 = User.objects.create_user(
            username="stud1", password="pw12345678", role=User.ROLE_STUDENT, first_name="Aman Raj"
        )
        StudentProfile.objects.create(user=self.s1, crn="101", course="CSE", semester=5, section="A")
        self.closed = AttendanceSession.objects.create(
            teacher=self.teacher, subject="AI", status=AttendanceSession.STATUS_CLOSED
        )
        Attendance.objects.create(student=self.s1, session=self.closed)

    def _auth(self, token):
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {token.key}")

    def test_csv_requires_teacher_role(self):
        self._auth(self.student_token)
        response = self.client.get(reverse("export-csv"))
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_csv_export_shape(self):
        self._auth(self.teacher_token)
        response = self.client.get(reverse("export-csv"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertIn("attachment", response["Content-Disposition"])

        content = response.content.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(content)))
        header = rows[0]
        self.assertEqual(header[0], "CRN")
        self.assertEqual(header[1], "Name")
        self.assertEqual(header[-1], "%")
        data_row = rows[1]
        self.assertEqual(data_row[0], "101")
        self.assertEqual(data_row[-1], "100.0")

    def test_csv_export_sanitizes_formula_like_names(self):
        risky_student = User.objects.create_user(
            username="stud2", password="pw12345678", role=User.ROLE_STUDENT, first_name="=cmd|'/c calc'"
        )
        StudentProfile.objects.create(user=risky_student, crn="=102", course="CSE", semester=5, section="A")
        Attendance.objects.create(student=risky_student, session=self.closed)

        self._auth(self.teacher_token)
        response = self.client.get(reverse("export-csv"))
        content = response.content.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(content)))
        risky_row = next(r for r in rows if r[0].endswith("102"))
        self.assertTrue(risky_row[0].startswith("'"))
        self.assertTrue(risky_row[1].startswith("'"))

    def test_excel_export_shape(self):
        from openpyxl import load_workbook
        import io as io_module

        self._auth(self.teacher_token)
        response = self.client.get(reverse("export-excel"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", response["Content-Disposition"])

        wb = load_workbook(io_module.BytesIO(response.content))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        self.assertEqual(header[0], "CRN")
        self.assertEqual(header[-1], "%")
        data_row = [cell.value for cell in ws[2]]
        self.assertEqual(data_row[0], "101")

    def test_excel_export_sanitizes_formula_like_names(self):
        from openpyxl import load_workbook
        import io as io_module

        risky_student = User.objects.create_user(
            username="stud3", password="pw12345678", role=User.ROLE_STUDENT, first_name="=cmd|'/c calc'"
        )
        StudentProfile.objects.create(user=risky_student, crn="=103", course="CSE", semester=5, section="A")
        Attendance.objects.create(student=risky_student, session=self.closed)

        self._auth(self.teacher_token)
        response = self.client.get(reverse("export-excel"))
        wb = load_workbook(io_module.BytesIO(response.content))
        ws = wb.active
        rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
        risky_row = next(r for r in rows if str(r[0]).endswith("103"))
        self.assertTrue(str(risky_row[0]).startswith("'"))
        self.assertTrue(str(risky_row[1]).startswith("'"))

    def test_pdf_export_shape(self):
        self._auth(self.teacher_token)
        response = self.client.get(reverse("export-pdf"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_pdf_export_handles_many_sessions_without_overflow(self):
        from pypdf import PdfReader
        import io as io_module

        for i in range(60):
            extra_session = AttendanceSession.objects.create(
                teacher=self.teacher, subject="AI", status=AttendanceSession.STATUS_CLOSED
            )
            Attendance.objects.create(student=self.s1, session=extra_session)

        self._auth(self.teacher_token)
        response = self.client.get(reverse("export-pdf"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        reader = PdfReader(io_module.BytesIO(response.content))
        self.assertGreaterEqual(len(reader.pages), 1)

    def test_pdf_column_widths_never_exceed_usable_width(self):
        from attendance.views import compute_pdf_column_widths

        for num_columns in (3, 10, 24, 60, 100, 200):
            widths = compute_pdf_column_widths(num_columns, usable_width=697.9)
            self.assertEqual(len(widths), num_columns)
            self.assertLessEqual(sum(widths), 697.9 + 1e-6)
            self.assertTrue(all(w >= 0 for w in widths))

    def test_pdf_export_strips_sanitizer_apostrophe(self):
        risky_student = User.objects.create_user(
            username="stud4", password="pw12345678", role=User.ROLE_STUDENT, first_name="=cmd|'/c calc'"
        )
        StudentProfile.objects.create(user=risky_student, crn="=104", course="CSE", semester=5, section="A")
        Attendance.objects.create(student=risky_student, session=self.closed)

        self._auth(self.teacher_token)
        response = self.client.get(reverse("export-pdf"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        # The raw sanitizer prefix ("'=104") should not appear verbatim in the PDF's
        # underlying byte stream content stream text-showing operators — this is a
        # coarse but effective smoke check without needing full PDF text extraction.
        self.assertNotIn(b"'=104", response.content)
